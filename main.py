#!/usr/bin/env python
# -*- coding: utf-8 -*- 

#
# GDG - Online Cloud Study Jams Vietnam - Quests Counter
# Author: GDG Cloud Hanoi
# Date: Aug 29, 2019
#

import argparse
import datetime
import io
import random
import shutil

import bs4
from console import fg, bg, fx, defx
import openpyxl
import requests

PROG_NAME = 'GDG Quest Counter'
DEBUG = False
COLORED_MODE = True

DATA = {
    'input_file': None,
    'participants': {},
    'result': {
        'error': None,
        'rank_by_location': {},
        'rank_by_timestamp': {},
    },
}

FILTER = {
    'skip_quests': ['GCP Essentials'],
    'date_range': [datetime.date(2019, 7, 28), datetime.date(2019, 8, 30)],
    'location': {
        'hanoi': {
            'title': 'Hà Nội',
            'names': ['hanoi', 'ha noi', 'hà nội'],
        },
        'danang': {
            'title': 'Đà Nẵng',
            'names': ['danang', 'da nang', 'đà nẵng'],
        },
        'hcm': {
            'title': 'Hồ Chí Minh City',
            'names': ['hcm', 'ho chi minh', 'hồ chí minh', 'thành phố hồ chí minh'],
        },
    }
}

OPTIONS = {
    'show_quest_detail': True,
    'process_top_people_only': 5,  # For testing only
    'hidden_email': True,
}

GDOCS_URL = {
    'file_id': '1VE2sH6zePhdwaSDir9ucUoXPYTXIjIR3eRFKQ-IVZcw',
    'sheet_id': '241580121',
    'format': 'xlsx',
    'template': 'https://docs.google.com/feeds/download/spreadsheets/Export' +
                '?key=%(file_id)s&exportFormat=%(format)s&gid=%(sheet_id)s',
}

DATE_FORMAT = '%b %d, %Y'
INDENT_LV1 = '    '
INDENT_LV2 = INDENT_LV1 + INDENT_LV1

STYLE_ERR = bg.lightred + fg.white
STYLE_WARN = bg.lightyellow + fg.blue
STYLE_INFO = bg.lightwhite + fg.black

STYLE_RANK_HEADER = bg.magenta + fg.white
STYLE_RANK_ACTIVE = bg.lightgreen + fg.blue
STYLE_RANK = bg.lightcyan + fg.blue
STYLE_RANK_INFO = bg.lightwhite + fg.black

# Styles
EOL = '\n'
COLOR_ALL = [ 'black', 'red', 'green', 'yellow', 'blue', 'magenta',
              'cyan', 'white', 'lightblack', 'lightred',
              'lightgreen', 'lightyellow', 'lightblue', 'lightmagenta',
              'lightcyan', 'lightwhite' ]

TERM_SIZE = shutil.get_terminal_size(fallback=(80, 30))

def cc(text, color):
    return color + text + fx.end if COLORED_MODE else text
    
def prt(*args):
    print(*args)
    
def prt_err(*args):
    prt(EOL, cc('ERROR', STYLE_ERR), *args)
    
def prt_warn(*args):
    prt(EOL, cc('WARNING', STYLE_WARN), *args)
    
def random_bg():
    while True:
        color = random.choice(COLOR_ALL)
        # BG must not be black?
        if color not in ('black', 'lightblack'):
            return color

def random_fg(bg=None):
    while True:
        color = random.choice(COLOR_ALL)
        # FG must not be same (or similar) as BG
        if not bg or not (color.endswith(bg) or bg.endswith(color)):
            return color

def parse_args():
    parser = argparse.ArgumentParser(prog=PROG_NAME)
    parser.add_argument('--debug', action='store_true',
        default=False,
        help='DEBUG mode')
    parser.add_argument('--no-color', dest='color_mode', action='store_false',
        default=True,
        help='NO COLOR mode')
    parser.add_argument('-i', '--input-file', dest='input_file',
        help='Input file of participants info (must be an Excel file)')
    parser.add_argument('--start-date', dest='start_date',
        default=datetime.datetime.strftime(FILTER['date_range'][0], '%Y-%m-%d'),
        help='Start date of quests to count (YYYY-MM-DD or none)')
    parser.add_argument('--end-date', dest='end_date',
        default=datetime.datetime.strftime(FILTER['date_range'][1], '%Y-%m-%d'),
        help='End date of quests to count (YYYY-MM-DD or none)')
    return parser.parse_args()

def main():
    global DEBUG, COLORED_MODE
    
    # Parse args
    args = parse_args()
    DEBUG = args.debug
    COLORED_MODE = args.color_mode
    
    # Date range of quests
    date_range = FILTER['date_range']
    start_date = args.start_date
    if start_date.lower() == 'none':
        date_range[0] = None
    else:
        date_range[0] = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = args.end_date
    if end_date.lower() == 'none':
        date_range[1] = None
    else:
        date_range[1] = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Input file
    input_file = args.input_file or download_input()
    DATA['input_file'] = input_file
    
    # Parse input
    parse_input(input_file)
    # Process input
    count_quests()
    
def download_input():
    prt(EOL,
        cc('DOWNLOADING GDOCS INPUT FILE ...', STYLE_INFO),
        EOL)
    
    filepath = 'result.%(format)s' % GDOCS_URL
    url = GDOCS_URL['template'] % GDOCS_URL
    req = requests.get(url)
    
    with open(filepath, 'w+b') as file:
        file.write(req.content)
    
    prt(INDENT_LV1 + 'Input file saved to', cc(filepath, fg.cyan), EOL)
    return filepath

def parse_input(input):
    wb = openpyxl.load_workbook(filename=input)
    sh = wb[wb.sheetnames[0]]
    
    participants = DATA['participants']
    rows_not_processed = []
    
    row_id = 0
    for row in sh.iter_rows():
        row_id += 1
        if row[0].value == 'Timestamp':
            # Skip header row
            pass
        elif not row[0].is_date:
            rows_not_processed.append(row)
        else:
            person = {
                'row_id': row_id,
                'timestamp': row[0].value,
                'email': row[1].value.strip().lower(),
                'name': row[2].value.strip(),
                'nick_name': row[3].value.strip(),
                'qwiklabs_link': row[4].value.strip(),
                'location': row[5].value.strip(),
                'quests': [],
                'legal_quests': [],
            }
            email = person['email']
            if email in participants:
                # Duplicated entry
                prt_warn('Dupplicated input entry ' + person['email'],
                         'at row %d' % row_id)
            participants[email] = person
    
    # Unprocessed rows
    show_unprocessed_rows(rows_not_processed)
        
def show_unprocessed_rows(rows):
    # TODO this should not happen
    if rows:
        prt(cc('IGNORED ROWS', STYLE_WARN)+ str(rows))
        prt(EOL, INDENT_LV1, str(rows))
    
def count_quests():
    participants = DATA['participants']
    person_index = 0
    process_only = OPTIONS['process_top_people_only'] if DEBUG else None
    for person in participants.values():
        try:
            count_quests_of(person)
        except Exception as ex:
            prt('Unable to parse QUESTS report for user %s' % person['email'])
        person_index += 1
        # For testing only
        if process_only and person_index == process_only:
            break
    
    # Track ERROR and OK reports
    error_list = []
    ok_list = []
    for person in participants.values():
        if person.get('error', None):
            error_list.append(person)
        else:
            ok_list.append(person)
    
    # Sort people by LEGAL quests
    ok_list.sort(key=lambda x: len(x['legal_quests']), reverse=True)
    
    # Filter and sort result by time submitting first quest
    def _pp_1st_quest_date_str(person):
        date = pp_1st_quest_date(person)
        return str(date) if date else 'z'  # Biggest string
    
    ok_list_by_time = list(ok_list)
    ok_list_by_time.sort(key=lambda p: _pp_1st_quest_date_str(p))
    
    rank_by_location = {
        'all': ok_list,
    }
    rank_by_timestamp = {
        'all': ok_list_by_time,
    }
    DATA['result'] = {
        'error': error_list,
        'rank_by_location': rank_by_location,
        'rank_by_timestamp': rank_by_timestamp,
    }
    
    # Filter and sort result by location/timestamp
    filter_loc = FILTER['location']
    for loc_name, loc_data in filter_loc.items():
        rank_by_location[loc_name] = []
        rank_by_timestamp[loc_name] = []
    
    for person in ok_list:
        person_loc = person['location'].lower()
        for loc_name, loc_data in filter_loc.items():
            if person_loc in loc_data['names']:
                rank_by_location[loc_name].append(person)
                
    for person in ok_list_by_time:
        person_loc = person['location'].lower()
        for loc_name, loc_data in filter_loc.items():
            if person_loc in loc_data['names']:
                rank_by_timestamp[loc_name].append(person)
                
    # Show final result on screen
    show_result_header(DATA['result'])
    # Errors
    show_result_error(error_list)
    # Result all
    show_result_by_loc('ALL LOCATIONS', ok_list)
    show_result_by_time(
        'ALL LOCATIONS (BY TIME SUBMITTING FIRST QUEST)',
        ok_list_by_time)
    # Result by location
    for loc_name, loc_data in filter_loc.items():
        show_result_by_loc(loc_data['title'].upper(), rank_by_location[loc_name])
        show_result_by_time(
            loc_data['title'].upper() + ' (BY TIME SUBMITTING FIRST QUEST)',
            rank_by_timestamp[loc_name])
    
    # Save result to text file
    save_result_txt()

def pp_1st_quest_date(person):
    quests = person['legal_quests']
    quest = quests[0] if len(quests) else None
    return quest['earned_date'] if quest else None
        
def show_email(person, txtmode=False):
    email = person['email']
    if OPTIONS['hidden_email']:
        email = '******@' + email.split('@')[1]
    return email if txtmode else cc(email, fg.cyan)

def show_result_header(result, outfile=None):
    if not outfile:
        prt(EOL, EOL,
            cc(INDENT_LV2+'HIGH SCORES'+INDENT_LV2, STYLE_INFO + fx.blink), EOL)
    else:
        outfile.write('\nGDG - CLOUD STUDY JAMS RESULT\n')
        all_participants = len(result['rank_by_location']['all'])
        outfile.write(INDENT_LV1 + 'Total participants: %d\n' % all_participants)
        
        filter_loc = FILTER['location']
        count_all_loc = 0
        for loc_name, loc_data in filter_loc.items():
            count = len(result['rank_by_location'][loc_name])
            count_all_loc += count
            outfile.write(INDENT_LV2 + loc_data['title'] + ': ' + str(count) + '\n')
        
        count_unknown = all_participants - count_all_loc
        outfile.write(INDENT_LV2 + 'Unknown location: %d\n' % count_unknown)
        
        outfile.write(INDENT_LV1 + 'Time period:\n')
        outfile.write(INDENT_LV2 + 'From Date: %s\n' % str(FILTER['date_range'][0]))
        outfile.write(INDENT_LV2 + 'To Date: %s\n' % str(FILTER['date_range'][1]))
        outfile.write('\n')

def show_result_error(plist, outfile=None):
    if not plist:
        return
        
    if not outfile:
        prt(EOL, cc('ERRORS', bg.lightred+fg.white+fx.bold))
        
        for person in plist:
            prt_err(EOL, person['name'], cc(person['error'], fg.red))
    else:
        ordinal = 0
        outfile.write('\n' + 'ERRORS' + '\n')
        for person in plist:
            ordinal += 1
            outfile.write('  %d. %s - %s\n' % (
                          ordinal, person['name'],
                          person['error']))

def show_result_by_loc(title, plist, outfile=None):
    if not outfile:
        prt(EOL, cc(" %s " % title, STYLE_RANK_HEADER), EOL)
        
        ordinal = 0
        for person in plist:
            ordinal += 1
            rank_style = STYLE_RANK_ACTIVE if ordinal <= 5 else STYLE_RANK
            prt(EOL, INDENT_LV1, cc('%2d.' % ordinal, rank_style),
                cc('%3d Legal ' % len(person['legal_quests']), STYLE_RANK_INFO),
                cc('%3d Total ' % len(person['quests']), STYLE_RANK_INFO),
                person['name'],)
        prt(EOL)
    else:
        ordinal = 0
        outfile.write('\n' + title + '\n')
        for person in plist:
            ordinal += 1
            outfile.write(INDENT_LV1 + '%d. %s - %d Legal quests (%d Total)\n' % (
                          ordinal, person['name'],
                          len(person['legal_quests']),
                          len(person['quests'])))

def show_result_by_time(title, plist, outfile=None):
    if not outfile:
        prt(EOL, cc(" %s " % title, STYLE_RANK_HEADER), EOL)
        
        ordinal = 0
        for person in plist:
            ordinal += 1
            rank_style = STYLE_RANK_ACTIVE if ordinal <= 5 else STYLE_RANK
            date = pp_1st_quest_date(person)
            date_str = str(date) if date else 'N/A'
            prt(EOL, INDENT_LV1,
                cc('%2d.' % ordinal, rank_style),
                cc(' Earliest Date %s ' % date_str, STYLE_RANK_INFO),
                person['name'])
        prt(EOL)
    else:
        ordinal = 0
        outfile.write('\n' + title + '\n')
        for person in plist:
            ordinal += 1
            outfile.write(INDENT_LV1 + '%d. %s - Earliest Date %s\n' % (
                          ordinal, person['name'],
                          str(pp_1st_quest_date(person))))

def save_result_txt():
    result = DATA['result']
    error_list = result['error']
    
    # Outfile for saving result
    with io.open('result.txt', 'w', encoding='utf-8') as outfile:
        # Header
        show_result_header(result, outfile=outfile)
        # Errors
        show_result_error(error_list, outfile=outfile)
        # Result all location
        show_result_by_loc('ALL LOCATIONS',
            result['rank_by_location']['all'], outfile=outfile)
        show_result_by_loc('ALL LOCATIONS (BY TIME SUBMITTING FIRST QUEST)',
            result['rank_by_timestamp']['all'], outfile=outfile)
        # Result per location
        filter_loc = FILTER['location']
        for loc_name, loc_data in filter_loc.items():
            show_result_by_loc(loc_data['title'].upper(),
                result['rank_by_location'][loc_name], outfile=outfile)
            show_result_by_time(loc_data['title'].upper() +
                ' (BY TIME SUBMITTING FIRST QUEST)',
                result['rank_by_timestamp'][loc_name], outfile=outfile)

    prt(EOL + cc('RESULT saved to ' + cc('result.txt', fg.cyan), STYLE_INFO))

def count_quests_of(person):
    qwiklabs_link = person['qwiklabs_link']
    resp = requests.get(qwiklabs_link)
    if resp.status_code != 200:  # Not OK
        prt_err('UNABLE to load QUESTS report for user %s' % person['email'])
        person['error'] = 'UNABLE to load QUESTS report page'
    else:
        html = bs4.BeautifulSoup(resp.content, features="html.parser")
        div_all_quests = html.body.find_all('div',
            attrs={'class': 'public-profile__badge'})
        if not div_all_quests:
            # Person has no quest complete
            pass
        else:
            quests_list = person['quests']
            # prt('Quest count = %d' % len(div_all_quests))
            for div in div_all_quests:
                child_tags = []
                for child in div.children:
                    if isinstance(child, bs4.element.Tag):
                        child_tags.append(child)
                if len(child_tags) != 3:
                    prt_err('UNEXPECTED quests report content')
                    person['error'] = 'UNEXPECTED quests report content'
                else:
                    title = child_tags[1].text.strip()
                    date_str = child_tags[2].text.strip().split('\n')[1]
                    date = datetime.datetime.strptime(date_str, DATE_FORMAT).date()
                    quest_info = {
                        'title': title,
                        'earned_date': date,
                    }
                    # prt(quest_info)
                    quests_list.append(quest_info)
            # prt(quests_list)
            show_quests_report_of(person)
            # Count legal quests
            legal_quests = []
            skip_quests = FILTER['skip_quests']
            from_date, to_date = FILTER['date_range']
            for quest in person['quests']:
                if (quest['title'] not in skip_quests and
                    (not from_date or from_date <= quest['earned_date']) and
                    (not to_date or to_date >= quest['earned_date'])):
                    legal_quests.append(quest)
            person['legal_quests'] = legal_quests

def show_quests_report_of(person):
    # Title line
    prt(EOL, cc('QUEST REPORT', bg.lightgreen+fg.yellow+fx.bold),
        person['name'], '(' + show_email(person) + ') -',
        str(len(person['quests'])), 'quests')
    # Show all quests
    console_width = int(TERM_SIZE[0] * 0.8)
    if OPTIONS['show_quest_detail']:
        est_line_len = 0
        quests_at_line = []
        for quest in person['quests']:
            title = quest['title']
            if len(title) + est_line_len <= console_width:
                est_line_len += len(title)
                quests_at_line.append(title)
            else:
                show_quests_at_line(person, quests_at_line)
                quests_at_line = [title]
                est_line_len = len(title)
        if len(quests_at_line):
            show_quests_at_line(person, quests_at_line)
        
def show_quests_at_line(person, quests_title):
    args = []
    for title in quests_title:
        title = (' %s ' % title) if COLORED_MODE else ('[%s]' % title)
        bgc = random_bg()
        fgc = random_fg(bgc)
        args.append(cc(title, getattr(bg, bgc) + getattr(fg, fgc)))
    prt(EOL+INDENT_LV1, *args)

if __name__ == '__main__':
    main()
